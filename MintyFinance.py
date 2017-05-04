"""MintyFinance.py.

This is the main script that will execute daily financial updates for the
MintyFinance .csv file. The main functions will come from minty.pu
"""
import sqlite3
import minty
import openpyxl

##############################################################################
# Retrieve financial data for today from MintAPI
today, sheet, day = minty.Timestamp()
accountTotals, categoryTotals = minty.MintyScrape(today)
accountBalances = minty.MintyBalance()
# print(sheet)
# print(today)
# print(accountTotals)
# print(categoryTotals)
# print(accountBalances)

###############################################################################
# SQLite3 Database for storing csv locations of each account type
connection = sqlite3.connect('accounts.db')
cursor = connection.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS ACCOUNTS
            (id_db INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            balance REAL );''')
for key, value in accountBalances.items():
    cursor.execute("INSERT or IGNORE INTO ACCOUNTS (name) VALUES (?)", [key])
    cursor.execute("UPDATE ACCOUNTS SET balance= ? WHERE name = ?",
                   (value, key))
connection.commit()

##############################################################################
# Retrieve all account values and store in query, will need to implement a
# more robust, reactive system to do this automatically in the future
query = []
for key in list(accountBalances.keys()):
    lookup = key
    cursor.execute("SELECT * FROM ACCOUNTS WHERE name = ?", (lookup,))
    row = cursor.fetchall()
    # print(query)
    query.append(row[0][2])
# print(query)

###############################################################################
# Read and write .CSV data for account balances
column = int(day) + 2
workbook = sheet + ".xlsx"
wb = openpyxl.load_workbook(workbook)
ws = wb[sheet]

ws.cell(row=24, column=column, value=query[2])  # Discover
ws.cell(row=25, column=column, value=query[0])  # Preferred
ws.cell(row=26, column=column, value=query[1])  # Sapphire
ws.cell(row=27, column=column, value=query[5])  # Checking
ws.cell(row=28, column=column, value=query[6])  # Savings
wb.save(workbook)

##############################################################################
# Need to write DB function and .csv function for categories now!
